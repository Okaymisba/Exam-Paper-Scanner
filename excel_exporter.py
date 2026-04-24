"""
Excel Exporter — OBE format
Columns are grouped by CLO with merged headers.
Structure:
  Row 1 : Title
  Row 2 : Roll No | Name | ←── CLO 1 ──→ | ←── CLO 2 ──→ | … | CLO Totals | Grand Total | % | Result
  Row 3 :            |    | Q1(/3)|Q2(/3) | Q3(/3)|Q4(/6) | … |            |             |   |
  Data  : one row per student
  Summary: average / highest / lowest
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
C_TITLE    = "1E3A5F"
C_CLO_HDR  = "2D6A9F"
C_Q_HDR    = "4A90D9"
C_TOT_HDR  = "1B5E20"
C_PASS_BG  = "D5F5E3"
C_FAIL_BG  = "FADBD8"
C_SUM_BG   = "EBF5FB"
C_ALT_BG   = "F4F6F7"
C_WHITE    = "FFFFFF"
C_PASS_FG  = "1D6A39"
C_FAIL_FG  = "C0392B"


def _fill(c):  return PatternFill(start_color=c, end_color=c, fill_type="solid")
def _font(bold=False, color=None, size=10, italic=False):
    kw = {"bold": bold, "size": size, "italic": italic}
    if color: kw["color"] = color
    return Font(**kw)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _centre():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():    return Alignment(horizontal="left",   vertical="center")


# ── Main export function ──────────────────────────────────────────────────────

def export_to_excel(output, students: list, setup: dict) -> None:
    exam_name      = setup.get("examName", "Exam")
    questions      = setup.get("questions", [])
    pass_threshold = float(setup.get("passThreshold", 50))

    # Build CLO structure
    clo_groups = defaultdict(list)   # clo_no → [question dicts]
    for q in questions:
        clo_groups[int(q["clo"])].append(q)
    clo_nos = sorted(clo_groups.keys())

    num_questions = len(questions)
    total_max     = sum(float(q.get("maxMarks", 0)) for q in questions)

    # Column layout:
    # A=Roll No, B=Name, C...(per question), then CLO totals, Grand Total, %, Result
    # We'll figure out column indices dynamically.

    # Fixed left columns
    COL_ROLL = 1
    COL_NAME = 2
    q_start  = 3                        # first question column
    q_end    = q_start + num_questions - 1

    clo_total_cols = {}                  # clo_no → col index
    col = q_end + 1
    for clo_no in clo_nos:
        clo_total_cols[clo_no] = col
        col += 1

    COL_GRAND_TOT = col;     col += 1
    COL_PCT       = col;     col += 1
    COL_RESULT    = col
    total_cols    = col

    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Marks"
    ws.sheet_view.showGridLines = False

    lc = get_column_letter  # shorthand

    # ── ROW 1 : Title ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{lc(total_cols)}1")
    c = ws["A1"]
    c.value     = f"{exam_name}  —  OBE Examination Marks"
    c.fill      = _fill(C_TITLE)
    c.font      = _font(bold=True, color=C_WHITE, size=14)
    c.alignment = _centre()
    ws.row_dimensions[1].height = 32

    # ── ROW 2 : Group headers (CLO spans + fixed columns) ────────────────────
    # Roll No — spans rows 2-3
    ws.merge_cells(start_row=2, start_column=COL_ROLL, end_row=3, end_column=COL_ROLL)
    _hdr(ws, 2, COL_ROLL, "Roll No", C_CLO_HDR)

    # Name — spans rows 2-3
    ws.merge_cells(start_row=2, start_column=COL_NAME, end_row=3, end_column=COL_NAME)
    _hdr(ws, 2, COL_NAME, "Student Name", C_CLO_HDR)

    # CLO group headers
    q_col = q_start
    for clo_no in clo_nos:
        qs = clo_groups[clo_no]
        clo_max = sum(float(q.get("maxMarks", 0)) for q in qs)
        span    = len(qs)
        if span > 1:
            ws.merge_cells(start_row=2, start_column=q_col,
                           end_row=2,   end_column=q_col + span - 1)
        _hdr(ws, 2, q_col,
             f"CLO {clo_no}  (max {int(clo_max)})", C_CLO_HDR)
        q_col += span

    # CLO total header columns — span rows 2-3
    for clo_no, col_idx in clo_total_cols.items():
        clo_max = sum(float(q.get("maxMarks", 0)) for q in clo_groups[clo_no])
        ws.merge_cells(start_row=2, start_column=col_idx,
                       end_row=3,   end_column=col_idx)
        _hdr(ws, 2, col_idx,
             f"CLO {clo_no}\nTotal\n(/{int(clo_max)})", C_TOT_HDR)

    # Grand Total, %, Result — span rows 2-3
    for col_idx, label in [
        (COL_GRAND_TOT, f"Grand\nTotal\n(/{int(total_max)})"),
        (COL_PCT,       "Percentage"),
        (COL_RESULT,    "Result"),
    ]:
        ws.merge_cells(start_row=2, start_column=col_idx,
                       end_row=3,   end_column=col_idx)
        _hdr(ws, 2, col_idx, label, C_TOT_HDR)

    ws.row_dimensions[2].height = 38

    # ── ROW 3 : Per-question sub-headers ─────────────────────────────────────
    q_col = q_start
    for clo_no in clo_nos:
        for q in clo_groups[clo_no]:
            q_no  = int(q.get("no", q_col - q_start + 1))
            q_max = float(q.get("maxMarks", 0))
            _hdr(ws, 3, q_col, f"Q{q_no}\n(/{int(q_max)})", C_Q_HDR)
            q_col += 1

    ws.row_dimensions[3].height = 34

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    # Build question-index → column mapping
    # questions list order is preserved, but we need to write them in CLO order.
    # The question columns are in CLO order (same as setup display).
    q_col_map = {}   # question index in `questions` list → excel column
    ec = q_start
    for clo_no in clo_nos:
        for q in clo_groups[clo_no]:
            # find index of this question in setup.questions
            q_no = int(q.get("no", 0))
            for i, orig_q in enumerate(questions):
                if int(orig_q.get("no", i + 1)) == q_no:
                    q_col_map[i] = ec
                    break
            ec += 1

    grand_totals = []

    for row_off, student in enumerate(students):
        r = row_off + 4   # data starts at row 4

        marks_list = student.get("marks", [])
        # Compute CLO subtotals and grand total
        clo_obtained = defaultdict(float)
        grand         = 0.0
        for idx, m in enumerate(marks_list):
            val = float(m.get("obtained", 0))
            grand += val
            clo_obtained[int(m.get("clo", 1))] += val

        pct    = (grand / total_max * 100) if total_max else 0
        result = "Pass" if pct >= pass_threshold else "Fail"
        grand_totals.append(grand)

        bg = C_PASS_BG if result == "Pass" else C_FAIL_BG

        def dc(col, value, bold=False, num_fmt=None, fg=None, align=None):
            cell = ws.cell(row=r, column=col, value=value)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=bold, color=fg, size=10)
            cell.border    = _border()
            cell.alignment = align or _centre()
            if num_fmt: cell.number_format = num_fmt
            return cell

        dc(COL_ROLL, student.get("rollNo", ""),  align=_centre())
        dc(COL_NAME, student.get("name",  ""),   align=_left())

        # Question marks
        for idx, m in enumerate(marks_list):
            col_idx = q_col_map.get(idx, q_start + idx)
            val     = float(m.get("obtained", 0))
            dc(col_idx, val if val != int(val) else int(val))

        # CLO totals
        for clo_no, col_idx in clo_total_cols.items():
            dc(col_idx, round(clo_obtained[clo_no], 1), bold=True)

        # Grand total, %, Result
        dc(COL_GRAND_TOT, round(grand, 1), bold=True)
        dc(COL_PCT, round(pct, 2), num_fmt='0.00"%"')
        rc = dc(COL_RESULT, result, bold=True,
                fg=C_PASS_FG if result == "Pass" else C_FAIL_FG)

    # ── SUMMARY ROWS ─────────────────────────────────────────────────────────
    if grand_totals:
        avg     = sum(grand_totals) / len(grand_totals)
        highest = max(grand_totals)
        lowest  = min(grand_totals)
        summary_data = [("Class Average", avg),
                        ("Highest Score", highest),
                        ("Lowest Score",  lowest)]

        sep_row = len(students) + 4
        ws.row_dimensions[sep_row].height = 6

        for j, (label, val) in enumerate(summary_data):
            r = sep_row + 1 + j
            for col in range(1, total_cols + 1):
                c = ws.cell(row=r, column=col)
                c.fill   = _fill(C_SUM_BG)
                c.border = _border()

            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r,   end_column=2)
            c = ws.cell(row=r, column=1, value=label)
            c.font      = _font(bold=True, size=10)
            c.fill      = _fill(C_SUM_BG)
            c.border    = _border()
            c.alignment = _centre()

            pct_val = val / total_max * 100 if total_max else 0
            c2 = ws.cell(row=r, column=COL_GRAND_TOT, value=round(val, 2))
            c2.font = _font(bold=True); c2.fill = _fill(C_SUM_BG)
            c2.border = _border(); c2.alignment = _centre()

            c3 = ws.cell(row=r, column=COL_PCT, value=round(pct_val, 2))
            c3.font = _font(bold=True); c3.fill = _fill(C_SUM_BG)
            c3.border = _border(); c3.alignment = _centre()
            c3.number_format = '0.00"%"'
            ws.row_dimensions[r].height = 22

    # ── COLUMN WIDTHS ─────────────────────────────────────────────────────────
    ws.column_dimensions[lc(COL_ROLL)].width = 13
    ws.column_dimensions[lc(COL_NAME)].width = 22
    for col in range(q_start, q_end + 1):
        ws.column_dimensions[lc(col)].width = 10
    for col in clo_total_cols.values():
        ws.column_dimensions[lc(col)].width = 12
    ws.column_dimensions[lc(COL_GRAND_TOT)].width = 12
    ws.column_dimensions[lc(COL_PCT)].width        = 12
    ws.column_dimensions[lc(COL_RESULT)].width     = 9

    ws.freeze_panes = "A4"
    wb.save(output)


# ── Helper for header cells ───────────────────────────────────────────────────

def _hdr(ws, row, col, label, bg_color):
    c = ws.cell(row=row, column=col, value=label)
    c.fill      = _fill(bg_color)
    c.font      = _font(bold=True, color=C_WHITE, size=10)
    c.alignment = _centre()
    c.border    = _border()
    return c
