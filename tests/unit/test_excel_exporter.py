"""
Unit tests for excel_exporter.export_to_excel.

These tests verify that the function writes a valid .xlsx workbook to the
output buffer and that the data rows, pass/fail classification, summary
rows, and edge cases all behave correctly without touching any external
service.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel_exporter import export_to_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_setup(exam_name="Test Exam", pass_threshold=50.0, questions=None):
    """Build a minimal setup dict for export_to_excel."""
    if questions is None:
        questions = [
            {"no": 1, "clo": 1, "maxMarks": 10},
            {"no": 2, "clo": 1, "maxMarks": 10},
            {"no": 3, "clo": 2, "maxMarks": 20},
        ]
    return {
        "examName": exam_name,
        "passThreshold": pass_threshold,
        "questions": questions,
    }


def _make_student(roll_no, obtained_marks):
    """Build a student dict with per-question marks."""
    marks = [
        {"questionNo": 1, "clo": 1, "maxMarks": 10, "obtained": obtained_marks[0]},
        {"questionNo": 2, "clo": 1, "maxMarks": 10, "obtained": obtained_marks[1]},
        {"questionNo": 3, "clo": 2, "maxMarks": 20, "obtained": obtained_marks[2]},
    ]
    return {"rollNo": roll_no, "name": f"Student {roll_no}", "marks": marks}


def _run_export(students, setup):
    """Run export_to_excel and return the resulting openpyxl Workbook."""
    buf = BytesIO()
    export_to_excel(buf, students, setup)
    buf.seek(0)
    return load_workbook(buf)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_export_produces_valid_xlsx_file():
    """Verify that export_to_excel writes a workbook that openpyxl can load."""
    setup = _make_setup()
    students = [_make_student("CS001", [8, 7, 15])]
    wb = _run_export(students, setup)
    assert wb is not None
    assert "Exam Marks" in wb.sheetnames


@pytest.mark.unit
def test_export_title_row_contains_exam_name():
    """The first cell of the title row should include the exam name."""
    setup = _make_setup(exam_name="Midterm 2024")
    students = [_make_student("CS001", [8, 7, 15])]
    wb = _run_export(students, setup)
    ws = wb.active
    title_value = ws.cell(row=1, column=1).value or ""
    assert "Midterm 2024" in title_value


@pytest.mark.unit
def test_export_student_roll_number_appears_in_data_rows():
    """Each student's roll number should appear in column A of data rows."""
    setup = _make_setup()
    students = [
        _make_student("CS001", [8, 7, 15]),
        _make_student("CS002", [3, 2, 5]),
    ]
    wb = _run_export(students, setup)
    ws = wb.active
    roll_values = [ws.cell(row=r, column=1).value for r in range(4, 6)]
    assert "CS001" in roll_values
    assert "CS002" in roll_values


@pytest.mark.unit
def test_export_pass_student_result_column():
    """A student scoring above the pass threshold should get 'Pass' in result column."""
    setup = _make_setup(pass_threshold=50.0)
    # 8+7+15 = 30 out of 40, pct = 75% => Pass
    students = [_make_student("CS001", [8, 7, 15])]
    wb = _run_export(students, setup)
    ws = wb.active
    # Result is the last data column; scan row 4 for "Pass"
    row_values = [ws.cell(row=4, column=c).value for c in range(1, 20)]
    assert "Pass" in row_values


@pytest.mark.unit
def test_export_fail_student_result_column():
    """A student scoring below the pass threshold should get 'Fail' in result column."""
    setup = _make_setup(pass_threshold=50.0)
    # 1+1+1 = 3 out of 40, pct = 7.5% => Fail
    students = [_make_student("CS009", [1, 1, 1])]
    wb = _run_export(students, setup)
    ws = wb.active
    row_values = [ws.cell(row=4, column=c).value for c in range(1, 20)]
    assert "Fail" in row_values


@pytest.mark.unit
def test_export_empty_student_list_does_not_raise():
    """Exporting with no students should complete without raising an exception."""
    setup = _make_setup()
    buf = BytesIO()
    export_to_excel(buf, [], setup)
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb is not None


@pytest.mark.unit
def test_export_summary_rows_present_when_students_exist():
    """When students are exported, summary rows (Average/Highest/Lowest) should appear."""
    setup = _make_setup()
    students = [
        _make_student("CS001", [8, 7, 15]),
        _make_student("CS002", [4, 3, 6]),
    ]
    wb = _run_export(students, setup)
    ws = wb.active
    # Summary rows start at row: len(students) + 4 + 1 = row 7 (sep) and 8, 9, 10
    all_values = []
    for row in ws.iter_rows(values_only=True):
        all_values.extend([str(v) for v in row if v is not None])
    assert any("Average" in v or "Highest" in v or "Lowest" in v for v in all_values)


@pytest.mark.unit
def test_export_custom_pass_threshold_boundary():
    """A student exactly at the pass threshold boundary should be marked Pass."""
    setup = _make_setup(pass_threshold=75.0)
    # 10+10+10 = 30 out of 40, pct = 75.0% => exactly at threshold => Pass
    students = [_make_student("CS001", [10, 10, 10])]
    wb = _run_export(students, setup)
    ws = wb.active
    row_values = [ws.cell(row=4, column=c).value for c in range(1, 20)]
    assert "Pass" in row_values


@pytest.mark.unit
def test_export_multiple_clo_groups_are_laid_out():
    """Questions from different CLOs should produce separate group headers."""
    setup = _make_setup(questions=[
        {"no": 1, "clo": 1, "maxMarks": 5},
        {"no": 2, "clo": 2, "maxMarks": 5},
        {"no": 3, "clo": 3, "maxMarks": 10},
    ])
    students = [{
        "rollNo": "CS001",
        "name": "Student CS001",
        "marks": [
            {"questionNo": 1, "clo": 1, "maxMarks": 5, "obtained": 4},
            {"questionNo": 2, "clo": 2, "maxMarks": 5, "obtained": 3},
            {"questionNo": 3, "clo": 3, "maxMarks": 10, "obtained": 7},
        ],
    }]
    # Should not raise; just verify three distinct CLO labels in row 2
    wb = _run_export(students, setup)
    ws = wb.active
    row2_values = [ws.cell(row=2, column=c).value for c in range(1, 15)]
    clo_headers = [v for v in row2_values if v and "CLO" in str(v)]
    assert len(clo_headers) >= 3
